# -*- coding: utf-8 -*-
"""
Created on Tue Jun 13 22:12:14 2023

@author: marca
"""


import pandas as pd
import numpy as np
import configparser
from sqlalchemy import create_engine, text
from sqlalchemy.exc import OperationalError
import os

# Get the directory of the main script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Get the full, system-independent path to the workspace
gpt_workspace = os.path.join(script_dir, "GPTworkspace")

def get_file_type(file_name):
    """Identify the type of the file based on its extension"""
    _, file_extension = os.path.splitext(file_name)
    return file_extension

def read_file(file_name):
    """Function to read a file based on its extension (.csv or .xlsx)"""
    file_type = get_file_type(file_name)
    file_path = os.path.join(gpt_workspace, file_name)

    if file_type == '.csv':
        return pd.read_csv(file_path)
    elif file_type == '.xlsx':
        return pd.read_excel(file_path)
    else:
        return f"Unsupported file type {file_type}."

def write_file(data, file_name):
    """Function to write to a file based on its extension (.csv or .xlsx)"""
    file_type = get_file_type(file_name)
    file_path = os.path.join(gpt_workspace, file_name)

    if file_type == '.csv':
        data.to_csv(file_path, index=False)
    elif file_type == '.xlsx':
        data.to_excel(file_path, index=False)
    else:
        return f"Unsupported file type {file_type}."

### TABULAR DATA TOOLS ###

def get_columns(file_name):
    """Get the columns of the file"""
    try:
        # Read the DataFrame using the helper function
        data = read_file(file_name)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Otherwise, get the columns
        columns = data.columns.tolist()
        return columns
    except Exception as e:
        return str(e)


def get_basic_table_info(filename):
    """Get basic information about the table in the file"""
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Get basic info
        info_dict = {
            'columns': list(data.columns),
            'num_rows': len(data),
            'num_columns': len(data.columns),
            'dtypes': dict(data.dtypes)
        }

        return info_dict
    except Exception as e:
        return str(e)

    
# Let's define the function "get_column_stats" that will take a filename and column name as input and return a dictionary with basic stats of the column.

def get_column_stats(filename, column_name):
    """Get the statistics of the given column in the file"""
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Check if the column exists in the DataFrame
        if column_name not in data.columns:
            return f"Column '{column_name}' not found in the file."

        # Check if the column is numeric or text-based
        if pd.api.types.is_numeric_dtype(data[column_name]):
            # Get basic stats for numeric columns
            stats = data[column_name].describe().to_dict()
        else:
            # Get basic stats for text-based columns
            stats = {
                'unique_values': data[column_name].nunique(),
                'most_frequent_value': data[column_name].mode().iloc[0] if not data[column_name].mode().empty else None,
                'freq_most_frequent_value': data[column_name].value_counts().iloc[0] if not data[column_name].value_counts().empty else None,
                'shortest_length': data[column_name].str.len().min(),
                'longest_length': data[column_name].str.len().max(),
                'average_length': data[column_name].str.len().mean()
            }

        return stats
    except Exception as e:
        return str(e)


def get_rows_by_index(filename, rows):
    """Get the rows with the given indices from the file"""
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Select only the rows with indices in the rows list
        selected_rows = data.iloc[rows]

        # Convert the selected rows back into CSV format
        csv_output = selected_rows.to_csv(index=False)

        return csv_output
    except Exception as e:
        return str(e)
    
def find_rows_by_value(filename, target_column, comparison_column, value):
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Find rows where the comparison column's value is equal to the given value
        result_rows = data[data[comparison_column] == value]

        # Return the values of the target column for these rows
        return result_rows[target_column]
    except Exception as e:
        return f"An error occurred: {e}"



def add_row_with_values(filename, row_values):
    """Add a row with the given values to the file"""
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Check if the number of values matches the number of columns
        if len(row_values) != len(data.columns):
            return "Error: The number of values provided does not match the number of columns in the file."

        # Append the new row
        data.loc[len(data)] = row_values

        # Write the DataFrame back to the file using the helper function
        write_file(data, filename)

        return f"Row with values {row_values} was added to '{filename}' successfully!"
    except Exception as e:
        return f"An error occurred: {e}"


def rename_columns(filename, old_names, new_names):
    """Rename the columns in the file"""
    # Check if both lists have the same length
    if len(old_names) != len(new_names):
        return "Error: Old names and new names lists must have the same length."

    # Create a dictionary from the lists of old and new names
    column_dict = dict(zip(old_names, new_names))

    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Rename the columns
        data.rename(columns=column_dict, inplace=True)

        # Write the DataFrame back to the file using the helper function
        write_file(data, filename)

        return f"Columns {old_names} were renamed to {new_names} in '{filename}' successfully!"
    except Exception as e:
        return f"An error occurred: {e}"




def read_file_sample(filename, n_rows=5):
    """Read the first n_rows of a file"""
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Slice the DataFrame to get only the first n_rows
        data = data.head(n_rows)

        # Convert the DataFrame to CSV text
        csv_text = data.to_csv(index=False)
        return csv_text
    except Exception as e:
        return f"An error occurred: {e}"


def delete_column(filename, column_name):
    """Delete a column from a file"""
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Check if the column exists in the DataFrame
        if column_name in data.columns:
            # Drop the column from the DataFrame
            data = data.drop(columns=[column_name])

            # Write the DataFrame back to the file using the helper function
            write_file(data, filename)

            return f"Column '{column_name}' deleted successfully!"
        else:
            return f"Column '{column_name}' not found in the file."
    except Exception as e:
        return f"An error occurred: {e}"

    
def add_column(filename, column_name):
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Add the new column to the DataFrame
        data[column_name] = pd.Series([np.nan]*len(data))

        # Write the DataFrame back to the file using the helper function
        write_file(data, filename)

        return f"Column '{column_name}' added successfully!"
    except Exception as e:
        return f"An error occurred: {e}"

def condense_table(filename, column_name):
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Check if the column exists in the DataFrame
        if column_name not in data.columns:
            return f"Column '{column_name}' not found in the file."

        # Process each group of rows that have the same target column value
        def process_group(group):
            # Detect the data types of the columns
            column_is_numeric = group.dtypes.apply(lambda dtype: np.issubdtype(dtype, np.number))

            # Sum numeric columns, concatenate string columns
            return group.agg({column: 'sum' if is_numeric else ' '.join for column, is_numeric in column_is_numeric.iteritems()})

        data = data.groupby(column_name).apply(process_group).reset_index()

        # Write the DataFrame back to the file using the helper function
        write_file(data, filename)

        return f"Table condensed successfully using column '{column_name}'!"
    except Exception as e:
        return f"An error occurred: {e}"

    
def drop_rows_by_condition(filename, column_name, condition):
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Check if the column exists in the DataFrame
        if column_name not in data.columns:
            return f"Column '{column_name}' not found in the file."

        # Apply the condition to the DataFrame
        # Invert the condition with '~' to keep rows where the condition is False
        data = data.query(f'~({column_name} {condition})')

        # Write the DataFrame back to the file using the helper function
        write_file(data, filename)

        return f"Rows dropped successfully using condition '{condition}' on column '{column_name}'!"
    except Exception as e:
        return f"An error occurred: {e}"

def populate_column_by_function(filename, target_columns, result_column, func_definition):
    # Create the global scope for the exec function
    global_scope = {"pd": pd, "np": np}

    # Use exec to define the function in the global scope
    exec(f"global func; func = {func_definition}", global_scope)

    # Get the function from the global scope
    func = global_scope["func"]

    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # If target_columns is specified
        if target_columns:
            # Check if the target_columns exist in the DataFrame
            missing_columns = [col for col in target_columns if col not in data.columns]
            if missing_columns:
                return f"Target columns '{missing_columns}' not found in the file."

            # Apply the function to the target columns to create the result column
            # If there's only one target column, we apply the function directly
            if len(target_columns) == 1:
                data[result_column] = data[target_columns[0]].apply(func)
            else:  # If there are multiple target columns, we need to apply function row-wise
                data[result_column] = data[target_columns].apply(lambda row: func(*row), axis=1)
        else:  # If no specific columns are specified, apply the function row-wise to the whole DataFrame
            data[result_column] = data.apply(func, axis=1)

        # Write the DataFrame back to the file using the helper function
        write_file(data, filename)

        return f"Column '{result_column}' created successfully using function '{func_definition}' on column(s) '{target_columns if target_columns else 'all columns'}'!"
    except Exception as e:
        return f"An error occurred: {e}"


def aggregate_dataframe_operations(filename, agg_operations):
    try:
        # Load the DataFrame using the helper function
        data = read_file(filename)

        # If it's a string, an error has occurred so return the error message
        if isinstance(data, str):
            return data

        # Executing the agg_operations
        global_scope = {'df': data, 'pd': pd, 'np': np}
        exec(f"df = df.{agg_operations}", global_scope)
        data = global_scope['df']

        # Write the DataFrame back to the file using the helper function
        output_filename = f"{os.path.splitext(filename)[0]}_pandas_working{get_file_type(filename)}"
        write_file(data, output_filename)

        return f"Aggregate operations completed successfully! The resulting DataFrame has been saved as '{output_filename}'."
    except Exception as e:
        return f"An error occurred: {e}"

    



### DATABASE TOOLS ###



def csv_to_sql(filename, table_name):
    # Parsing the configuration file
    config = configparser.ConfigParser()
    config.read('config.ini')
    
    # Ensure the database type is in lowercase
    database_type = config.get('database', 'type').lower()
    
    db_config = config[database_type]

    # Create the appropriate connection string
    if database_type == 'sqlite':
        conn_str = f"{database_type}:///{db_config['dbname']}.db"
    else:
        conn_str = f"{database_type}://{db_config['user']}:{db_config['password']}@{db_config['host']}:{db_config['port']}/{db_config['dbname']}"

    # Create a SQLAlchemy engine
    try:
        engine = create_engine(conn_str)
    except OperationalError:
        return f"Unable to create the engine due to an Operational Error. Please check your connection details and try again."

    filepath = os.path.join(gpt_workspace, filename)
    try:
        # Load the DataFrame from the CSV file
        data = pd.read_csv(filepath)

        # Save the DataFrame to the SQL database
        data.to_sql(table_name, engine, if_exists='replace')

        return f"CSV file '{filename}' was saved as table '{table_name}' in the {database_type} database successfully!"
    except Exception as e:
        return f"An error occurred: {e}"



def sql_to_csv(table_name, filename):
    # Parsing the configuration file
    config = configparser.ConfigParser()
    config.read('config.ini')

    # Getting the database type from the configuration
    database_type = config.get('database', 'type').lower()

    # Setting up the engine string based on the database type
    if database_type == 'sqlite':
        engine_string = f'sqlite:///{config.get("sqlite", "dbname")}.db'
    elif database_type == 'mysql':
        engine_string = f'mysql+pymysql://{config.get("mysql", "user")}:{config.get("mysql", "password")}@{config.get("mysql", "host")}:{config.get("mysql", "port")}/{config.get("mysql", "dbname")}'
    elif database_type == 'postgresql':
        engine_string = f'postgresql://{config.get("postgresql", "user")}:{config.get("postgresql", "password")}@{config.get("postgresql", "host")}:{config.get("postgresql", "port")}/{config.get("postgresql", "dbname")}'

    try:
        # Creating the engine
        engine = create_engine(engine_string)

        # Reading the data from the SQL table into a DataFrame
        data = pd.read_sql_table(table_name, engine)

        # Writing the DataFrame to a CSV file in the gpt_workspace
        filepath = os.path.join(gpt_workspace, filename)
        data.to_csv(filepath, index=False)

        # Closing the connection
        engine.dispose()

        return f"Table '{table_name}' from the database has been successfully saved as '{filename}' in the 'gpt_workspace'!"
    except Exception as e:
        return f"An error occurred: {e}"


# =============================================================================
# def execute_sql(table_name, sql):
#     # Parsing the configuration file
#     config = configparser.ConfigParser()
#     config.read('config.ini')
# 
#     # Getting the database type from the configuration
#     database_type = config.get('database', 'type').lower()
# 
#     # Setting up the engine string based on the database type
#     if database_type == 'sqlite':
#         engine_string = f'sqlite:///{config.get("sqlite", "dbname")}.db'
#     elif database_type == 'mysql':
#         engine_string = f'mysql+pymysql://{config.get("mysql", "user")}:{config.get("mysql", "password")}@{config.get("mysql", "host")}:{config.get("mysql", "port")}/{config.get("mysql", "dbname")}'
#     elif database_type == 'postgresql':
#         engine_string = f'postgresql://{config.get("postgresql", "user")}:{config.get("postgresql", "password")}@{config.get("postgresql", "host")}:{config.get("postgresql", "port")}/{config.get("postgresql", "dbname")}'
# 
#     try:
#         # Creating the engine
#         engine = create_engine(engine_string)
# 
#         # Executing the SQL block
#         result = engine.execute(sql)
#         
# 
#         # Closing the connection
#         engine.dispose()
# 
#         # Handling the results
#         if result.returns_rows:
#             return [dict(row) for row in result]
#         else:
#             return f"SQL block executed successfully on table '{table_name}'!"
#     except Exception as e:
#         print(e)
#         return f"An error occurred: {e}"
# =============================================================================
    
    
def execute_sql(table_name, sql):
    # Parsing the configuration file
    config = configparser.ConfigParser()
    config.read('config.ini')

    # Getting the database type from the configuration
    database_type = config.get('database', 'type').lower()

    # Setting up the engine string based on the database type
    if database_type == 'sqlite':
        engine_string = f'sqlite:///{config.get("sqlite", "dbname")}.db'
    elif database_type == 'mysql':
        engine_string = f'mysql+pymysql://{config.get("mysql", "user")}:{config.get("mysql", "password")}@{config.get("mysql", "host")}:{config.get("mysql", "port")}/{config.get("mysql", "dbname")}'
    elif database_type == 'postgresql':
        engine_string = f'postgresql://{config.get("postgresql", "user")}:{config.get("postgresql", "password")}@{config.get("postgresql", "host")}:{config.get("postgresql", "port")}/{config.get("postgresql", "dbname")}'

    try:
        # Creating the engine
        engine = create_engine(engine_string)

        # Executing the SQL block
        with engine.connect() as connection:
            result = connection.execute(text(sql))
            
            # Handling the results
            if result.returns_rows:
                # Convert the result to a list of dictionaries
                columns = result.keys()
                return [dict(zip(columns, row)) for row in result]
            else:
                return f"SQL block executed successfully on table '{table_name}'!"
    except Exception as e:
        print(e)
        return f"An error occurred: {e}"


### EXCEL ONLY TOOLS ###

from openpyxl import load_workbook
from openpyxl.styles import Font, Color

def change_font_excel_cells(filename, cell_positions, font_name, font_size, font_color):
    try:
        # Load the Excel Workbook
        file_path = os.path.join(gpt_workspace, filename)
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            return f"File {filename} does not exist."

        # Get the active worksheet
        ws = wb.active

        # Create a Font object
        font = Font(name=font_name, size=font_size, color=Color(rgb=font_color))

        # Apply the font to the cells
        for cell_position in cell_positions:
            cell = ws[cell_position]
            cell.font = font

        # Save the workbook
        wb.save(file_path)

        return f"Font and color changed successfully for cells {cell_positions} in file '{filename}'!"
    except Exception as e:
        return f"An error occurred: {e}"

    

def adjust_column_width_excel(filename, columns, widths):
    try:
        # Load the Excel Workbook
        file_path = os.path.join(gpt_workspace, filename)
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            return f"File {filename} does not exist."

        # Get the active worksheet
        ws = wb.active

        # Ensure columns and widths lists are of equal length, or if single width is provided for all
        if isinstance(widths, list):
            if len(columns) != len(widths):
                return f"Error: Length of columns list ({len(columns)}) doesn't match length of widths list ({len(widths)})."
        else:
            widths = [widths]*len(columns)

        # Apply the width to the columns
        for column, width in zip(columns, widths):
            ws.column_dimensions[column].width = width

        # Save the workbook
        wb.save(file_path)

        return f"Width adjusted successfully for columns {columns} in file '{filename}'!"
    except Exception as e:
        return f"An error occurred: {e}"



def get_cell_info_excel(filename, cell_positions):
    try:
        # Load the Excel Workbook
        file_path = os.path.join(gpt_workspace, filename)
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            return f"File {filename} does not exist."

        # Get the active worksheet
        ws = wb.active

        # Initialize an empty dictionary for cell information
        cell_info = {}

        # Fetch the info for each cell
        for cell_position in cell_positions:
            cell = ws[cell_position]
            
            # Store cell info in a dictionary
            cell_info[cell_position] = {
                'font': cell.font.name,
                'font_size': cell.font.sz,
                'font_color': cell.font.color.rgb,
                'width': ws.column_dimensions[cell.column_letter].width,
                'height': ws.row_dimensions[cell.row].height
            }

        return cell_info
    except Exception as e:
        return f"An error occurred: {e}"



